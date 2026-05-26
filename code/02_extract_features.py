"""
Sexual Assault Opinion NLP Analysis Pipeline
============================================
Analyzes judicial opinions for:
  - How victims of sexual assault are referred to and framed
  - Agency patterns (who acts, who is acted upon)
  - Hedging and epistemic distancing
  - Blame attribution language
  - Victim credibility challenges

Usage:
    python sexual_assault_nlp_pipeline.py <path_to_docx_or_txt> [case_name]

Outputs:
    - <case_name>_annotated.xlsx  — sentence-level annotated spreadsheet
    - <case_name>_stats.json      — aggregate statistics for cross-case comparison
"""

import re
import sys
import json
import os
from pathlib import Path
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# PATTERN DICTIONARIES
# Each key maps to a list of compiled regex patterns.
# ─────────────────────────────────────────────────────────────────────────────

# How victims/plaintiffs are referred to
VICTIM_REFERENCE_PATTERNS = {
    "formal_legal_neutral": [
        re.compile(r'\b[Pp]laintiff\b'),
        re.compile(r'\b[Pp]etitioner\b'),
        re.compile(r'\b[Cc]laimant\b'),
    ],
    "name_reference": [
        re.compile(r'\bGatewood\b'),
        re.compile(r'\bRoscoe\b'),
    ],
    "victim_label": [
        re.compile(r'\b[Vv]ictim\b'),
        re.compile(r'\b[Ss]urvivor\b'),
        re.compile(r'\b[Cc]omplainant\b'),
    ],
    "gendered_pronoun": [
        re.compile(r'\b(he|him|his|she|her|hers)\b', re.IGNORECASE),
    ],
    "subject_position": [],  # computed during analysis
}

# How perpetrators/defendants are referred to
PERP_REFERENCE_PATTERNS = {
    "formal_legal": [
        re.compile(r'\b[Dd]efendant\b'),
        re.compile(r'\b[Rr]espondent\b'),
    ],
    "name_reference": [
        re.compile(r'\bSeidel\b'),
        re.compile(r'\bMindy\b'),
        re.compile(r'\bTekton\b'),
    ],
    "title_honorific": [
        re.compile(r'\bMs\.\s+Seidel\b'),
        re.compile(r'\bMr\.\s+\w+\b'),
    ],
}

# Hedging / epistemic distancing language
HEDGING_PATTERNS = {
    "allegation_framing": [
        re.compile(r'\b[Aa]lleg\w+\b'),          # alleged, alleges, allegation
        re.compile(r'\b[Cc]laim\w*\b'),           # claims, claimed
        re.compile(r'\b[Aa]ssert\w*\b'),          # asserts, assertion
        re.compile(r'\b[Cc]ontend\w*\b'),
        re.compile(r'\b[Pp]urport\w*\b'),
        re.compile(r'\b[Ss]o.call\w*\b'),
        re.compile(r'\b[Aa]ccus\w*\b'),
    ],
    "modal_hedging": [
        re.compile(r'\b[Mm]ay have\b'),
        re.compile(r'\b[Mm]ight have\b'),
        re.compile(r'\b[Cc]ould have\b'),
        re.compile(r'\b[Ww]ould have\b'),
        re.compile(r'\b[Aa]ppear\w*\b'),
        re.compile(r'\b[Ss]eem\w*\b'),
    ],
    "scare_quotes": [
        re.compile(r'"[^"]{1,40}"'),              # quoted terms (short)
        re.compile(r"'[^']{1,40}'"),
    ],
    "plausibility_framing": [
        re.compile(r'\b[Pp]lausibl\w+\b'),
        re.compile(r'\b[Ff]acially\b'),
        re.compile(r'\b[Aa]t this stage\b'),
        re.compile(r'\b[Tt]aken as true\b'),
        re.compile(r'\b[Aa]ssuming\b'),
    ],
}

# Agency patterns — who performs actions on whom
AGENCY_PATTERNS = {
    "victim_active_agency": [
        re.compile(r'\b[Pp]laintiff\b.{0,30}\b(filed|brought|alleged|reported|complained|initiated|sought|asserted|argued|presented|invoked)\b'),
        re.compile(r'\bGatewood\b.{0,30}\b(filed|brought|alleged|reported|complained|initiated|sought|asserted|argued)\b'),
    ],
    "victim_passive_object": [
        re.compile(r'\b(assault\w*|harass\w*|touch\w*|victimiz\w*|abus\w*)\b.{0,50}\b[Pp]laintiff\b'),
        re.compile(r'\b[Pp]laintiff\b.{0,30}\b(was|were)\b.{0,30}\b(assault\w*|harass\w*|touch\w*|abus\w*)\b'),
    ],
    "perp_active_agency": [
        re.compile(r'\bSeidel\b.{0,40}\b(assault\w*|harass\w*|touch\w*|obtain\w*|utiliz\w*|employ\w*|access\w*)\b'),
        re.compile(r'\bDefendant\b.{0,40}\b(assault\w*|harass\w*|touch\w*|obtain\w*|utiliz\w*|fail\w*)\b'),
    ],
    "perp_passive_object": [
        re.compile(r'\b(accused|named|sued|charged)\b.{0,40}\bSeidel\b'),
        re.compile(r'\bSeidel\b.{0,30}\bwas\b.{0,30}\b(accused|named)\b'),
    ],
    "institutional_agency": [
        re.compile(r'\b[Cc]ourt\b.{0,40}\b(dismissed|granted|denied|held|found|concluded|ordered)\b'),
        re.compile(r'\b[Tt]ekton\b.{0,40}\b(failed|argued|moved|sought|contend\w*)\b'),
    ],
}

# Blame attribution patterns
BLAME_PATTERNS = {
    "victim_blame_direct": [
        re.compile(r'\b[Pp]laintiff\b.{0,60}\b(fail\w*|neglect\w*|delay\w*|did not|didn.t|should have|could have)\b'),
        re.compile(r'\b(fail\w*|neglect\w*|should have)\b.{0,60}\b[Pp]laintiff\b'),
        re.compile(r'\bPlaintiff failed to respond\b'),
        re.compile(r'\bPlaintiff.{0,20}did not present\b'),
    ],
    "victim_responsibility": [
        re.compile(r'\b(knew|know|known|aware)\b.{0,60}\b(Plaintiff|Gatewood)\b'),
        re.compile(r'\b(Plaintiff|Gatewood)\b.{0,60}\b(knew|know|known|was aware|should have known|in the exercise of reasonable diligence)\b'),
        re.compile(r'\breasonable diligence\b'),
        re.compile(r'\bshould have known\b'),
    ],
    "perp_blame": [
        re.compile(r'\bSeidel\b.{0,60}\b(assault\w*|harass\w*|violat\w*|misus\w*|inappropriat\w*)\b'),
        re.compile(r'\bTekton\b.{0,60}\b(fail\w*|neglect\w*|inadequat\w*)\b'),
    ],
    "procedural_deflection": [
        re.compile(r'\bstatute of limitations\b'),
        re.compile(r'\btime.barred\b'),
        re.compile(r'\bbarred by\b'),
        re.compile(r'\blimitations period\b'),
        re.compile(r'\bdismissed\b.{0,40}\b(time|limitations|barred)\b'),
        re.compile(r'\bfailure to (state|respond|plead)\b'),
    ],
    "credibility_challenge": [
        re.compile(r'\b(inconsistent|inconsistency|contradictory|changed|removed|omitted|deliberately)\b'),
        re.compile(r'\b(implausible|impossible|factually impossible)\b'),
        re.compile(r'\bperfunctory\b'),
        re.compile(r'\bno coherent argument\b'),
        re.compile(r'\bnot persuasive\b'),
    ],
}

# Assault terminology — how the violence itself is named
ASSAULT_TERMINOLOGY = {
    "clinical_legal": [
        re.compile(r'\b[Ss]exual [Aa]ssault\b'),
        re.compile(r'\b[Ss]exual [Bb]attery\b'),
        re.compile(r'\b[Ss]exual [Hh]arassment\b'),
        re.compile(r'\b[Mm]isconduct\b'),
    ],
    "minimizing_terms": [
        re.compile(r'\b[Ii]nappropriately touching\b'),
        re.compile(r'\b[Ss]pending more time\b'),
        re.compile(r'\b[Bb]reach of policy\b'),
        re.compile(r'\b[Pp]rotocol\b'),
    ],
    "direct_violence_terms": [
        re.compile(r'\b[Aa]ssault\w*\b'),
        re.compile(r'\b[Bb]attery\b'),
        re.compile(r'\b[Aa]bus\w*\b'),
        re.compile(r'\b[Vv]iolat\w*\b'),
    ],
}

# Gendered pronoun tracking
PRONOUN_PATTERNS = {
    "masculine": re.compile(r'\b(he|him|his)\b', re.IGNORECASE),
    "feminine": re.compile(r'\b(she|her|hers)\b', re.IGNORECASE),
    "neutral": re.compile(r'\b(they|them|their)\b', re.IGNORECASE),
}

# Potential gendered errors (wrong pronoun for party)
GENDERED_ERROR_PATTERNS = [
    # Male plaintiff referred to as "her" or "she"
    re.compile(r'\b[Pp]laintiff.{0,30}\b(she|her)\b'),
    re.compile(r'\b(she|her)\b.{0,30}\b[Pp]laintiff\b'),
    re.compile(r'\b(his|her)\b.{0,20}(third|response|reply|brief)\b'),
]


# ─────────────────────────────────────────────────────────────────────────────
# TEXT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_text(filepath: str) -> str:
    """Extract text from .docx or .txt file."""
    path = Path(filepath)
    if path.suffix.lower() in ('.docx', '.doc'):
        from docx import Document
        doc = Document(filepath)
        return '\n'.join(p.text for p in doc.paragraphs)
    elif path.suffix.lower() == '.txt':
        return path.read_text(encoding='utf-8')
    elif path.suffix.lower() == '.pdf':
        # Basic fallback — users should convert PDFs first for best results
        raise ValueError("PDF extraction not supported in this version. Convert to .docx or .txt first.")
    else:
        # Try reading as plain text
        return path.read_text(encoding='utf-8', errors='replace')


def extract_metadata(text: str) -> dict:
    """Extract basic case metadata from header."""
    metadata = {}

    # Case name
    m = re.search(r'^(.+?)\n', text.strip())
    if m:
        metadata['case_name'] = m.group(1).strip()

    # Court
    m = re.search(r'(United States .+? Court.+?)\n', text)
    if m:
        metadata['court'] = m.group(1).strip()

    # Date decided
    m = re.search(r'(\w+ \d+, \d{4}), Decided', text)
    if m:
        metadata['date_decided'] = m.group(1)

    # Case number
    m = re.search(r'Case No\.\s+(\S+)', text)
    if m:
        metadata['case_number'] = m.group(1)

    # Judge
    m = re.search(r'Opinion by:\s*(.+)', text)
    if m:
        metadata['judge'] = m.group(1).strip()

    return metadata


# ─────────────────────────────────────────────────────────────────────────────
# SENTENCE SPLITTING (regex-based, no NLTK needed)
# ─────────────────────────────────────────────────────────────────────────────

def split_sentences(text: str) -> list:
    """
    Split legal text into sentences using regex.
    Handles abbreviations common in legal text (e.g., "U.S.", "§", "v.", "Inc.").
    """
    # Remove footnote markers like [*1], [*2]
    text = re.sub(r'\[\*\d+\]', '', text)
    # Normalize whitespace
    text = re.sub(r'\n{2,}', '\n\n', text)
    text = re.sub(r'[ \t]+', ' ', text)

    # Common legal abbreviations that should NOT end sentences
    abbrevs = r'(?:U\.S\.|S\.Ct\.|F\.\d+d|F\.Supp\.|§|Art\.|Sec\.|No\.|v\.|et al\.|Inc\.|LLC\.|Ltd\.|Corp\.|viz\.|i\.e\.|e\.g\.|cf\.|id\.|supra\.|infra\.|pp\.|p\.|vol\.|ed\.|vs\.|Mr\.|Mrs\.|Ms\.|Dr\.|Prof\.|Jan\.|Feb\.|Mar\.|Apr\.|Jun\.|Jul\.|Aug\.|Sep\.|Oct\.|Nov\.|Dec\.)'

    # Replace abbreviation periods temporarily
    text = re.sub(abbrevs, lambda m: m.group(0).replace('.', '§§'), text)

    # Split on sentence-ending punctuation
    sentences = re.split(r'(?<=[.!?])\s+(?=[A-Z"])', text)

    # Restore abbreviation periods
    sentences = [s.replace('§§', '.').strip() for s in sentences if s.strip()]

    return sentences


# ─────────────────────────────────────────────────────────────────────────────
# CORE ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

def analyze_sentence(sentence: str, sentence_idx: int) -> dict:
    """Run all pattern analyses on a single sentence. Returns annotation dict."""

    result = {
        "sentence_id": sentence_idx + 1,
        "sentence": sentence,
        "word_count": len(sentence.split()),

        # Victim reference
        "victim_ref_formal": 0,
        "victim_ref_name": 0,
        "victim_label": 0,
        "victim_pronoun_correct": 0,  # correct pronoun (he/him/his for male victim)
        "victim_pronoun_wrong_gender": 0,  # wrong pronoun (she/her for male victim)

        # Perpetrator reference
        "perp_ref_formal": 0,
        "perp_ref_name": 0,
        "perp_honorific": 0,

        # Hedging
        "hedging_allegation": 0,
        "hedging_modal": 0,
        "hedging_scare_quotes": 0,
        "hedging_plausibility": 0,
        "hedging_total": 0,

        # Agency
        "victim_active_subject": 0,
        "victim_passive_object": 0,
        "perp_active_subject": 0,
        "perp_passive_object": 0,
        "institutional_agency": 0,

        # Blame
        "victim_blame_direct": 0,
        "victim_responsibility": 0,
        "perp_blame": 0,
        "procedural_deflection": 0,
        "credibility_challenge": 0,

        # Assault terminology
        "clinical_legal_term": 0,
        "minimizing_term": 0,
        "direct_violence_term": 0,

        # Tags (human-readable summary)
        "tags": [],
        "flags": [],
        "analyst_notes": "",
    }

    # ── Victim references ──────────────────────────────────────────
    for p in VICTIM_REFERENCE_PATTERNS["formal_legal_neutral"]:
        if p.search(sentence):
            result["victim_ref_formal"] += 1

    for p in VICTIM_REFERENCE_PATTERNS["name_reference"]:
        if p.search(sentence):
            result["victim_ref_name"] += 1

    for p in VICTIM_REFERENCE_PATTERNS["victim_label"]:
        if p.search(sentence):
            result["victim_label"] += 1
            result["tags"].append("victim_label")

    # Pronoun tracking (victim is male in this case — he/him/his are correct)
    if PRONOUN_PATTERNS["masculine"].search(sentence) and (
        re.search(r'\b[Pp]laintiff\b', sentence) or re.search(r'\bGatewood\b', sentence)
    ):
        result["victim_pronoun_correct"] += 1

    # Check for female pronoun applied near Plaintiff (gendered error)
    for p in GENDERED_ERROR_PATTERNS:
        if p.search(sentence):
            result["victim_pronoun_wrong_gender"] += 1
            result["flags"].append("GENDER_PRONOUN_ERROR: female pronoun used for male plaintiff")

    # Check "her" appearing in contexts that seem to reference plaintiff's filings
    if re.search(r'\bher\b.{0,30}(response|brief|motion|filing)\b', sentence, re.IGNORECASE):
        if not re.search(r'\bSeidel\b|\bMs\b', sentence):
            result["flags"].append("POSSIBLE_GENDER_ERROR: 'her' may incorrectly reference male plaintiff's filing")

    # ── Perpetrator references ─────────────────────────────────────
    for p in PERP_REFERENCE_PATTERNS["formal_legal"]:
        if p.search(sentence):
            result["perp_ref_formal"] += 1

    for p in PERP_REFERENCE_PATTERNS["name_reference"]:
        if p.search(sentence):
            result["perp_ref_name"] += 1

    for p in PERP_REFERENCE_PATTERNS["title_honorific"]:
        if p.search(sentence):
            result["perp_honorific"] += 1
            result["tags"].append("perp_honorific_title")

    # ── Hedging ───────────────────────────────────────────────────
    for p in HEDGING_PATTERNS["allegation_framing"]:
        if p.search(sentence):
            result["hedging_allegation"] += len(p.findall(sentence))

    for p in HEDGING_PATTERNS["modal_hedging"]:
        if p.search(sentence):
            result["hedging_modal"] += len(p.findall(sentence))

    for p in HEDGING_PATTERNS["scare_quotes"]:
        if p.search(sentence):
            result["hedging_scare_quotes"] += len(p.findall(sentence))
            result["tags"].append("scare_quotes")

    for p in HEDGING_PATTERNS["plausibility_framing"]:
        if p.search(sentence):
            result["hedging_plausibility"] += len(p.findall(sentence))

    result["hedging_total"] = (
        result["hedging_allegation"] + result["hedging_modal"] +
        result["hedging_scare_quotes"] + result["hedging_plausibility"]
    )
    if result["hedging_total"] > 0:
        result["tags"].append("hedging")

    # ── Agency ────────────────────────────────────────────────────
    for p in AGENCY_PATTERNS["victim_active_agency"]:
        if p.search(sentence):
            result["victim_active_subject"] += 1
            result["tags"].append("victim_active_subject")

    for p in AGENCY_PATTERNS["victim_passive_object"]:
        if p.search(sentence):
            result["victim_passive_object"] += 1
            result["tags"].append("victim_passive_object")

    for p in AGENCY_PATTERNS["perp_active_agency"]:
        if p.search(sentence):
            result["perp_active_subject"] += 1
            result["tags"].append("perp_active_subject")

    for p in AGENCY_PATTERNS["perp_passive_object"]:
        if p.search(sentence):
            result["perp_passive_object"] += 1

    for p in AGENCY_PATTERNS["institutional_agency"]:
        if p.search(sentence):
            result["institutional_agency"] += 1
            result["tags"].append("institutional_agency")

    # ── Blame ─────────────────────────────────────────────────────
    for p in BLAME_PATTERNS["victim_blame_direct"]:
        if p.search(sentence):
            result["victim_blame_direct"] += 1
            result["tags"].append("victim_blame_direct")
            result["flags"].append("VICTIM BLAME: direct attribution of fault/failure to victim")

    for p in BLAME_PATTERNS["victim_responsibility"]:
        if p.search(sentence):
            result["victim_responsibility"] += 1
            result["tags"].append("victim_responsibility")

    for p in BLAME_PATTERNS["perp_blame"]:
        if p.search(sentence):
            result["perp_blame"] += 1
            result["tags"].append("perp_blame_present")

    for p in BLAME_PATTERNS["procedural_deflection"]:
        if p.search(sentence):
            result["procedural_deflection"] += 1
            result["tags"].append("procedural_deflection")

    for p in BLAME_PATTERNS["credibility_challenge"]:
        if p.search(sentence):
            result["credibility_challenge"] += 1
            result["tags"].append("credibility_challenge")
            result["flags"].append("CREDIBILITY CHALLENGE: victim's account implicitly or explicitly questioned")

    # ── Assault terminology ────────────────────────────────────────
    for p in ASSAULT_TERMINOLOGY["clinical_legal"]:
        if p.search(sentence):
            result["clinical_legal_term"] += 1
            result["tags"].append("assault_clinical_term")

    for p in ASSAULT_TERMINOLOGY["minimizing_terms"]:
        if p.search(sentence):
            result["minimizing_term"] += 1
            result["tags"].append("assault_minimizing_term")
            result["flags"].append("MINIMIZING LANGUAGE: violence described in minimizing/bureaucratic terms")

    for p in ASSAULT_TERMINOLOGY["direct_violence_terms"]:
        if p.search(sentence):
            result["direct_violence_term"] += 1

    # ── Deduplicate tags ──────────────────────────────────────────
    result["tags"] = sorted(set(result["tags"]))
    result["flags"] = sorted(set(result["flags"]))

    return result


def analyze_document(text: str, case_name: str = "unknown") -> dict:
    """Full document analysis. Returns metadata + sentence-level results + aggregates."""

    metadata = extract_metadata(text)
    metadata['case_name_label'] = case_name

    sentences = split_sentences(text)
    results = [analyze_sentence(s, i) for i, s in enumerate(sentences)]

    # ── Aggregate statistics ───────────────────────────────────────
    agg = defaultdict(int)
    flagged_sentences = []
    tag_counter = Counter()

    for r in results:
        for key in r:
            if isinstance(r[key], int) and key not in ('sentence_id', 'word_count'):
                agg[key] += r[key]
        for tag in r['tags']:
            tag_counter[tag] += 1
        if r['flags']:
            flagged_sentences.append(r)

    total_sentences = len(results)
    total_words = sum(r['word_count'] for r in results)

    # Ratios
    agg_ratios = {}
    for key in ('hedging_total', 'victim_blame_direct', 'victim_responsibility',
                 'procedural_deflection', 'credibility_challenge',
                 'victim_active_subject', 'victim_passive_object',
                 'perp_active_subject', 'minimizing_term', 'perp_blame'):
        agg_ratios[f'{key}_per_100_sentences'] = round(
            agg[key] / total_sentences * 100, 2) if total_sentences else 0

    # Agency balance: victim active vs passive
    victim_active = agg['victim_active_subject']
    victim_passive = agg['victim_passive_object']
    agency_ratio = victim_active / max(victim_passive, 1)

    # Blame balance: victim blame vs perpetrator blame
    victim_blame_total = agg['victim_blame_direct'] + agg['victim_responsibility']
    perp_blame_total = agg['perp_blame']
    blame_ratio = victim_blame_total / max(perp_blame_total, 1)

    stats = {
        'case_name': case_name,
        'metadata': metadata,
        'total_sentences': total_sentences,
        'total_words': total_words,
        'aggregate_counts': dict(agg),
        'aggregate_ratios': agg_ratios,
        'tag_frequency': dict(tag_counter.most_common()),
        'flagged_sentence_count': len(flagged_sentences),
        'agency_ratio_victim_active_to_passive': round(agency_ratio, 3),
        'blame_ratio_victim_to_perp': round(blame_ratio, 3),
        'hedging_density': round(agg['hedging_total'] / total_sentences, 3) if total_sentences else 0,
        'sentence_results': results,
        'flagged_sentences': flagged_sentences,
    }

    return stats


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

# Color palette for tag categories
COLORS = {
    "victim_blame": "FFD7D7",      # soft red
    "credibility_challenge": "FFBCBC",  # stronger red
    "hedging": "FFF3CC",           # yellow
    "victim_passive": "E8D5F5",    # light purple
    "perp_active": "D5EAF5",       # light blue
    "minimizing": "FFE5CC",        # orange
    "procedural": "D5F5D5",        # green
    "pronoun_error": "FF6666",     # bright red
    "default": "FFFFFF",
}


def get_row_color(row_data: dict) -> str:
    if row_data.get('victim_pronoun_wrong_gender', 0) > 0:
        return COLORS["pronoun_error"]
    if row_data.get('victim_blame_direct', 0) > 0 or row_data.get('credibility_challenge', 0) > 0:
        return COLORS["victim_blame"]
    if row_data.get('minimizing_term', 0) > 0:
        return COLORS["minimizing"]
    if row_data.get('hedging_total', 0) > 3:
        return COLORS["hedging"]
    if row_data.get('victim_passive_object', 0) > 0:
        return COLORS["victim_passive"]
    if row_data.get('perp_active_subject', 0) > 0:
        return COLORS["perp_active"]
    if row_data.get('procedural_deflection', 0) > 0:
        return COLORS["procedural"]
    return COLORS["default"]


def write_excel(stats: dict, output_path: str):
    """Write full annotated Excel workbook."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, size=13, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D3A4A")
    subheader_fill = PatternFill("solid", fgColor="4A6FA5")
    subheader_font = Font(bold=True, color="FFFFFF")

    ws_summary.append(["SEXUAL ASSAULT OPINION ANALYSIS"])
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.append([f"Case: {stats['case_name']}"])
    ws_summary.append([f"Court: {stats['metadata'].get('court', 'N/A')}"])
    ws_summary.append([f"Date: {stats['metadata'].get('date_decided', 'N/A')}"])
    ws_summary.append([f"Judge: {stats['metadata'].get('judge', 'N/A')}"])
    ws_summary.append([f"Case No: {stats['metadata'].get('case_number', 'N/A')}"])
    ws_summary.append([])

    summary_data = [
        ["DOCUMENT STATISTICS", ""],
        ["Total sentences analyzed", stats['total_sentences']],
        ["Total words", stats['total_words']],
        ["Flagged sentences", stats['flagged_sentence_count']],
        [],
        ["AGENCY ANALYSIS", ""],
        ["Victim as active subject (sentences)", stats['aggregate_counts'].get('victim_active_subject', 0)],
        ["Victim as passive object (sentences)", stats['aggregate_counts'].get('victim_passive_object', 0)],
        ["Perpetrator as active subject (sentences)", stats['aggregate_counts'].get('perp_active_subject', 0)],
        ["Agency ratio (victim active:passive)", stats['agency_ratio_victim_active_to_passive']],
        ["  → ratio < 1.0 = victim more passive than active", ""],
        [],
        ["BLAME ATTRIBUTION", ""],
        ["Sentences with direct victim blame", stats['aggregate_counts'].get('victim_blame_direct', 0)],
        ["Sentences with victim-responsibility framing", stats['aggregate_counts'].get('victim_responsibility', 0)],
        ["Sentences with perpetrator blame", stats['aggregate_counts'].get('perp_blame', 0)],
        ["Blame ratio (victim:perp)", stats['blame_ratio_victim_to_perp']],
        ["  → ratio > 1.0 = more blame assigned to victim than perp", ""],
        [],
        ["HEDGING & EPISTEMIC DISTANCING", ""],
        ["Total hedging instances", stats['aggregate_counts'].get('hedging_total', 0)],
        ["Hedging density (per sentence)", stats['hedging_density']],
        ["Allegation-framing terms", stats['aggregate_counts'].get('hedging_allegation', 0)],
        ["Modal hedges (may, might, could)", stats['aggregate_counts'].get('hedging_modal', 0)],
        ["Scare quote instances", stats['aggregate_counts'].get('hedging_scare_quotes', 0)],
        [],
        ["ASSAULT TERMINOLOGY", ""],
        ["Clinical/legal terms (sexual assault, battery, etc.)", stats['aggregate_counts'].get('clinical_legal_term', 0)],
        ["Minimizing terms (inappropriately touching, etc.)", stats['aggregate_counts'].get('minimizing_term', 0)],
        ["Direct violence terms (assault, abuse, violate)", stats['aggregate_counts'].get('direct_violence_term', 0)],
        [],
        ["CREDIBILITY & LANGUAGE", ""],
        ["Credibility challenge instances", stats['aggregate_counts'].get('credibility_challenge', 0)],
        ["Procedural deflection sentences", stats['aggregate_counts'].get('procedural_deflection', 0)],
        ["Pronoun gender errors", stats['aggregate_counts'].get('victim_pronoun_wrong_gender', 0)],
    ]

    for row in summary_data:
        ws_summary.append(row)
        if row and len(row) == 2 and row[1] == "":
            # Section header
            cell = ws_summary.cell(row=ws_summary.max_row, column=1)
            cell.font = subheader_font
            cell.fill = subheader_fill

    ws_summary.column_dimensions['A'].width = 52
    ws_summary.column_dimensions['B'].width = 18

    # ── Sheet 2: Annotated Sentences ──────────────────────────────
    ws_ann = wb.create_sheet("Annotated Sentences")

    columns = [
        "ID", "Sentence", "Words",
        "Victim Ref\n(Formal)", "Victim Ref\n(Name)", "Victim Label",
        "Pronoun\nCorrect", "Pronoun\nGender Error",
        "Perp Ref\n(Formal)", "Perp Ref\n(Name)", "Perp\nHonorific",
        "Hedging\nTotal", "Hedge:\nAllegation", "Hedge:\nModal", "Hedge:\nScare Quotes",
        "Victim\nActive Subj", "Victim\nPassive Obj",
        "Perp\nActive Subj", "Institutional\nAgency",
        "Victim\nBlame Direct", "Victim\nResponsibility",
        "Perp\nBlame", "Procedural\nDeflection", "Credibility\nChallenge",
        "Clinical\nTerm", "Minimizing\nTerm", "Direct\nViolence Term",
        "Tags", "FLAGS"
    ]

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws_ann.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor="2D3A4A")
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    ws_ann.row_dimensions[1].height = 40

    # Data rows
    for row_data in stats['sentence_results']:
        row_color = get_row_color(row_data)
        fill = PatternFill("solid", fgColor=row_color)

        values = [
            row_data['sentence_id'],
            row_data['sentence'],
            row_data['word_count'],
            row_data['victim_ref_formal'],
            row_data['victim_ref_name'],
            row_data['victim_label'],
            row_data['victim_pronoun_correct'],
            row_data['victim_pronoun_wrong_gender'],
            row_data['perp_ref_formal'],
            row_data['perp_ref_name'],
            row_data['perp_honorific'],
            row_data['hedging_total'],
            row_data['hedging_allegation'],
            row_data['hedging_modal'],
            row_data['hedging_scare_quotes'],
            row_data['victim_active_subject'],
            row_data['victim_passive_object'],
            row_data['perp_active_subject'],
            row_data['institutional_agency'],
            row_data['victim_blame_direct'],
            row_data['victim_responsibility'],
            row_data['perp_blame'],
            row_data['procedural_deflection'],
            row_data['credibility_challenge'],
            row_data['clinical_legal_term'],
            row_data['minimizing_term'],
            row_data['direct_violence_term'],
            ', '.join(row_data['tags']),
            ' | '.join(row_data['flags']),
        ]

        ws_ann.append(values)
        current_row = ws_ann.max_row
        for col_idx in range(1, len(columns) + 1):
            cell = ws_ann.cell(row=current_row, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(size=9)
            if col_idx in (20, 21, 22, 24, 29):  # highlight important cols
                if cell.value and str(cell.value) not in ('0', ''):
                    cell.font = Font(size=9, bold=True, color="CC0000")

    # Column widths
    ws_ann.column_dimensions['A'].width = 5
    ws_ann.column_dimensions['B'].width = 70
    ws_ann.column_dimensions['C'].width = 7
    for col_letter in [get_column_letter(i) for i in range(4, len(columns))]:
        ws_ann.column_dimensions[col_letter].width = 10
    ws_ann.column_dimensions[get_column_letter(len(columns))].width = 45
    ws_ann.column_dimensions[get_column_letter(len(columns)-1)].width = 30

    ws_ann.freeze_panes = 'C2'

    # ── Sheet 3: Flagged Sentences ─────────────────────────────────
    ws_flag = wb.create_sheet("Flagged Sentences")

    flag_cols = ["ID", "Sentence", "FLAGS", "Tags"]
    for col_idx, col_name in enumerate(flag_cols, 1):
        cell = ws_flag.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8B0000")
        cell.alignment = Alignment(wrap_text=True, horizontal='center')

    for row_data in stats['flagged_sentences']:
        ws_flag.append([
            row_data['sentence_id'],
            row_data['sentence'],
            ' | '.join(row_data['flags']),
            ', '.join(row_data['tags']),
        ])
        current_row = ws_flag.max_row
        fill = PatternFill("solid", fgColor="FFD7D7")
        for col_idx in range(1, 5):
            cell = ws_flag.cell(row=current_row, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws_flag.column_dimensions['A'].width = 5
    ws_flag.column_dimensions['B'].width = 80
    ws_flag.column_dimensions['C'].width = 55
    ws_flag.column_dimensions['D'].width = 35

    # ── Sheet 4: Tag Frequency ─────────────────────────────────────
    ws_tags = wb.create_sheet("Tag Frequency")
    ws_tags.append(["Tag", "Sentence Count", "% of All Sentences"])
    ws_tags['A1'].font = Font(bold=True)
    ws_tags['B1'].font = Font(bold=True)
    ws_tags['C1'].font = Font(bold=True)

    for tag, count in sorted(stats['tag_frequency'].items(), key=lambda x: -x[1]):
        pct = round(count / stats['total_sentences'] * 100, 1) if stats['total_sentences'] else 0
        ws_tags.append([tag, count, f"{pct}%"])

    ws_tags.column_dimensions['A'].width = 35
    ws_tags.column_dimensions['B'].width = 18
    ws_tags.column_dimensions['C'].width = 20

    wb.save(output_path)
    print(f"Excel saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def analyze_file(filepath: str, case_name: str = None, output_dir: str = ".") -> dict:
    """Analyze a single file. Returns stats dict and writes Excel + JSON."""
    if case_name is None:
        case_name = Path(filepath).stem[:40]

    safe_name = re.sub(r'[^\w\-]', '_', case_name)

    print(f"\n{'='*60}")
    print(f"Analyzing: {case_name}")
    print(f"File: {filepath}")
    print(f"{'='*60}")

    text = extract_text(filepath)
    stats = analyze_document(text, case_name)

    print(f"Sentences: {stats['total_sentences']}")
    print(f"Words: {stats['total_words']}")
    print(f"Flagged sentences: {stats['flagged_sentence_count']}")
    print(f"Agency ratio (victim active:passive): {stats['agency_ratio_victim_active_to_passive']}")
    print(f"Blame ratio (victim:perp): {stats['blame_ratio_victim_to_perp']}")
    print(f"Hedging density: {stats['hedging_density']}")

    # Write Excel
    excel_path = os.path.join(output_dir, f"{safe_name}_annotated.xlsx")
    write_excel(stats, excel_path)

    # Write JSON stats (for cross-case aggregation)
    json_stats = {k: v for k, v in stats.items() if k != 'sentence_results'}
    json_path = os.path.join(output_dir, f"{safe_name}_stats.json")
    with open(json_path, 'w') as f:
        json.dump(json_stats, f, indent=2)
    print(f"Stats JSON saved: {json_path}")

    return stats


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python sexual_assault_nlp_pipeline.py <filepath> [case_name] [output_dir]")
        sys.exit(1)

    filepath = sys.argv[1]
    case_name = sys.argv[2] if len(sys.argv) > 2 else None
    output_dir = sys.argv[3] if len(sys.argv) > 3 else "."

    analyze_file(filepath, case_name, output_dir)
